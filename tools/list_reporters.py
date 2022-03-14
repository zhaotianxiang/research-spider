import os
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from utils import read_csv, download_img


def get_reporter_by_id(filepath):
    reports = {}
    for idx, line in enumerate(read_csv(filepath)):
        if idx == 0:
            continue
        if reports.get(line[4]):
            info = reports[line[4]]

            info["article_cnt"] += 1
            info["name"] = line[5]

            if line[6] != "":
                if line[6] not in info["email_list"]:
                    info["email_list"].append(line[6])

            if line[8] != "":
                if line[8] not in info["image_urls"]:
                    info["image_urls"].append(line[8])

            if line[11] != "":
                if line[11] not in info["twitter"]:
                    info["twitter"].append(line[11])

            if line[12] != "":
                if line[12] not in info["facebook"]:
                    info["twitter"].append(line[12])
        else:
            info = {}
            info["name"] = line[5]
            info["job_name"] = line[7]
            if line[6].strip():
                info["email_list"] = [line[6]]
            else:
                info["email_list"] = []

            if line[8] != "":
                info["image_urls"] = [line[8]]
            else:
                info["image_urls"] = []

            info["twitter"] = [line[11]]
            info["facebook"] = [line[12]]
            info["article_cnt"] = 1

            reports[line[4]] = info
    return reports


def download_reporter_image(reporters):
    if not os.path.exists("img"):
        os.mkdir("img")

    for key, value in tqdm(reporters.items()):
        img_filename = "{}.jpg".format(str(key))
        img_filepath = os.path.join("img", img_filename)
        if len(value["image_urls"]) == 0:
            continue
        url = value["image_urls"][0]
        download_img(url, img_filepath)
    print("照片下载成功")


def create_reporter_xls(filepath, reporters):
    wb = Workbook()
    sh = wb.active
    sh.row_dimensions[1].height = 20

    head = ["图片", "姓名", "职位", "邮箱", "twitter", "facebook", "简介", "发文数量"]
    for j in range(1, 9):
        sh.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
        sh.cell(1, j).value = head[j-1]

    sh.column_dimensions['A'].width = 13
    sh.column_dimensions['B'].width = 9
    sh.column_dimensions['C'].width = 9
    sh.column_dimensions['D'].width = 20
    sh.column_dimensions['E'].width = 13
    sh.column_dimensions['F'].width = 13
    sh.column_dimensions['G'].width = 25
    sh.column_dimensions['H'].width = 9

    for idx, (key, value) in enumerate(reporters.items()):

        row = idx + 2
        sh.row_dimensions[row].height = 100     # 设置行高

        if key != "99999":
            img_filepath = "./img/{}.jpg".format(str(key))
            img = Image(img_filepath)
            img.width, img.height = 100, 130
            sh.add_image(img, "A{}".format(str(row)))

        for j in range(1, 9):
            sh.cell(row, j).alignment = Alignment(horizontal='center', vertical='center')
        sh.cell(row, 2).value = value["name"]
        sh.cell(row, 3).value = value["job_name"]
        sh.cell(row, 4).value = value["email_list"][0] if len(value["email_list"]) else ""
        sh.cell(row, 5).value = value["twitter"][0]
        sh.cell(row, 6).value = value["facebook"][0]
        sh.cell(row, 8).value = value["article_cnt"]

    wb.save(filepath)


if __name__ == '__main__':

    reporters = get_reporter_by_id(filepath="data/kbs.csv")
    # reporters = pickle_load(filepath="reporters-byId.pkl")                # 数据源
    # reporters = pickle_dump(reporters, filepath="reporters-byId.pkl")       # 数据源
    # download_reporter_image(reporters)
    create_reporter_xls("reporters.xls", reporters=reporters)
