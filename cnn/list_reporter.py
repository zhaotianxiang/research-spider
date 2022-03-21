import csv
from tools.utils import read_csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image


def create_reporter_xls(filepath, reporters):
    wb = Workbook()
    sh = wb.active
    sh.row_dimensions[1].height = 20

    head = ["图片", "id", "姓名", "简介", "码值列表", "url", "图片URL"]
    for j in range(1, 8):
        sh.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
        sh.cell(1, j).value = head[j-1]

    sh.column_dimensions['A'].width = 19
    sh.column_dimensions['B'].width = 9
    sh.column_dimensions['C'].width = 9
    sh.column_dimensions['D'].width = 20
    sh.column_dimensions['E'].width = 13
    sh.column_dimensions['F'].width = 13
    sh.column_dimensions['G'].width = 25

    for idx, line in enumerate(reporters):
        if idx == 0:
            continue

        row = idx + 1
        sh.row_dimensions[row].height = 100     # 设置行高

        filename = line[5].split("/")[-1]
        img_filepath = "imges/" + filename
        img_filepath = img_filepath

        img = Image(img_filepath)
        img.width, img.height = 150, 130
        sh.add_image(img, "A{}".format(str(row)))

        for j in range(1, 8):
            sh.cell(row, j).alignment = Alignment(horizontal='center', vertical='center')
            if j == 1:
                continue
            else:
                sh.cell(row, j).value = line[j-2]

    wb.save(filepath)


if __name__ == '__main__':
    lines = read_csv(filepath="./data/csv/cnn-reporter.csv")
    create_reporter_xls(filepath="./data/csv/cnn-reporter.xls", reporters=lines)