import os
import json
from tqdm import tqdm
from utils import download_img
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image


def txt_2json(filepath):
    """txt中每行保存一条dict数据，为rocket查询结果"""
    with open(filepath, "r", encoding="utf8") as fr:
        for line in fr.readlines():
            dict0 = json.loads(line)
            yield dict0


def get_info(dict0):
    """将rocketreach返回的response进行解析，生成特定字段数据"""
    if not dict0.get("people"):
        return
    peoples = dict0.get("people")

    res = []
    for people in peoples:

        userid = people["id"]
        name = people["name"]
        img_url = people["profile_pic"]             # 头像URL

        if not os.path.exists("img"):
            os.mkdir("img")
        img_path = os.path.join("img", "{}.jpg".format(userid))
        if not os.path.isfile(img_path):
            download_img(url=img_url, img_filepath=img_path)

        links = "\n".join(["{}:{}".format(key, value) for key, value in people["links"].items()])       # 个人社交账号

        url = people["url"]                         # rocket，个人url
        location = people["location"]               # 定位
        title = people["current_title"]             # 头衔
        company = people["current_employer"]        # 公司

        if people.get("teaser"):
            teaser = people.get("teaser")
            emails = "\n".join(teaser["emails"])
            phones = "\n".join([item.get("number") for item in teaser["phones"]])
        else:
            emails = ""
            phones = ""

        numbers = emails + "\n" + phones
        res.append([userid, name, company, title, numbers, links, location, url, img_url])
    return res


def create_rocket_xls(filepath, items):
    wb = Workbook()
    sh = wb.active
    sh.row_dimensions[1].height = 20

    # head字段与items中的数据保持对齐
    head = ["图片", "姓名", "公司", "职位", "码值", "社交账号", "位置", "url", "头像url"]

    # 初始化抬头
    for j in range(1, len(head)+1):
        sh.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
        sh.cell(1, j).value = head[j-1]

    sh.column_dimensions['A'].width = 13
    for i in range(1, len(head)+1):                     # 设置列宽
        sh.column_dimensions[chr(65+i)].width = 18

    for idx, (item) in enumerate(items):

        row = idx + 2
        sh.row_dimensions[row].height = 100             # 设置行高

        img_filepath = "./img/{}.jpg".format(str(item[0]))
        img = Image(img_filepath)
        img.width, img.height = 100, 130
        sh.add_image(img, "A{}".format(str(row)))

        for j in range(2, len(head)+1):                 # 单元格格式初始化:水平居中、垂直居中
            sh.cell(row, j).alignment = Alignment(horizontal='center', vertical='center')

            sh.cell(row, j).value = item[j-1]

    wb.save(filepath)


if __name__ == '__main__':
    dicts = txt_2json(filepath="input.txt")

    all_info = []
    for dict0 in tqdm(dicts):
        info = get_info(dict0)
        all_info.extend(info)

    create_rocket_xls(filepath="rocket.xls", items=all_info)



